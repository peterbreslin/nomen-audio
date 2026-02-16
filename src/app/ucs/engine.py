"""UCS engine — parses UCS 8.2.1 spreadsheets into lookup tables."""

from dataclasses import dataclass

import openpyxl


@dataclass(frozen=True)
class CatInfo:
    """Full details for a single UCS CatID entry."""

    cat_id: str
    category: str
    subcategory: str
    cat_short: str
    explanation: str
    synonyms: list[str]

    @property
    def category_full(self) -> str:
        return f"{self.category}-{self.subcategory}"


# ---------------------------------------------------------------------------
# Module-level singletons (populated by load_ucs)
# ---------------------------------------------------------------------------

_categories: list[str] = []
_subcategories: dict[str, list[str]] = {}
_catid_to_info: dict[str, CatInfo] = {}
_cat_sub_to_catid: dict[tuple[str, str], str] = {}
_category_explanations: dict[str, str] = {}
_synonym_index: dict[str, list[str]] = {}
_loaded: bool = False


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_ucs(full_path: str, top_path: str) -> None:
    """Parse both UCS spreadsheets and populate all lookup tables."""
    global _loaded
    _parse_full_list(full_path)
    _parse_top_level(top_path)
    _build_synonym_index()
    _loaded = True


def _parse_full_list(path: str) -> None:
    """Parse 'UCS v8.2.1' sheet: rows 4+ with headers at row 3."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["UCS v8.2.1"]

    cats_set: set[str] = set()
    subs: dict[str, list[str]] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        category = row[0]
        if category is None:
            break

        subcategory = str(row[1] or "")
        cat_id = str(row[2] or "")
        cat_short = str(row[3] or "")
        explanation = str(row[4] or "")
        synonyms_raw = str(row[5] or "")

        synonyms = _parse_synonyms(synonyms_raw)

        info = CatInfo(
            cat_id=cat_id,
            category=str(category),
            subcategory=subcategory,
            cat_short=cat_short,
            explanation=explanation,
            synonyms=synonyms,
        )
        _catid_to_info[cat_id] = info
        _cat_sub_to_catid[(str(category), subcategory)] = cat_id

        cats_set.add(str(category))
        subs.setdefault(str(category), []).append(subcategory)

    wb.close()

    _categories.clear()
    _categories.extend(sorted(cats_set))

    _subcategories.clear()
    _subcategories.update(subs)


def _parse_top_level(path: str) -> None:
    """Parse 'USC Category Overview' sheet: rows 3+ with headers at row 2."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["USC Category Overview"]

    for row in ws.iter_rows(min_row=3, values_only=True):
        category = row[0]
        if category is None:
            break
        explanation = str(row[1] or "")
        _category_explanations[str(category)] = explanation

    wb.close()


def _parse_synonyms(raw: str) -> list[str]:
    """Split comma-separated synonym string into a clean list."""
    if not raw or raw == "None":
        return []
    return [s.strip() for s in raw.split(",") if s.strip()]


# Extra synonyms not in the official UCS spreadsheet.
# Keyed by CatID, values are lowercase synonyms to inject.
_EXTRA_SYNONYMS: dict[str, list[str]] = {
    "GUNCano": ["cannon"],
}


def _build_synonym_index() -> None:
    """Build reverse index: lowercase synonym → list of CatIDs."""
    _synonym_index.clear()
    for cat_id, info in _catid_to_info.items():
        for syn in info.synonyms:
            key = syn.lower()
            _synonym_index.setdefault(key, []).append(cat_id)
    for cat_id, extras in _EXTRA_SYNONYMS.items():
        for syn in extras:
            _synonym_index.setdefault(syn.lower(), []).append(cat_id)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def is_loaded() -> bool:
    return _loaded


def get_categories() -> list[str]:
    return list(_categories)


def get_subcategories(category: str) -> list[str]:
    return list(_subcategories.get(category, []))


def get_catid_info(cat_id: str) -> CatInfo | None:
    return _catid_to_info.get(cat_id)


def lookup_catid(category: str, subcategory: str) -> str | None:
    return _cat_sub_to_catid.get((category, subcategory))


def get_category_explanation(category: str) -> str | None:
    return _category_explanations.get(category)


def get_synonyms(cat_id: str) -> list[str]:
    info = _catid_to_info.get(cat_id)
    return list(info.synonyms) if info else []


def get_synonym_index() -> dict[str, list[str]]:
    return _synonym_index


def get_all_catinfo() -> list[CatInfo]:
    """Return all 753 CatInfo entries from the lookup table."""
    return list(_catid_to_info.values())
